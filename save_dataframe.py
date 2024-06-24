from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



def save_dataframe(app, study_type, gen_info: list, all_devices: list,
                   setting_report: dict, detailed_fls: list):
    """ saves the dataframe in the user directory.
    If the user is connected through citrix, the file should
    be saved local users PowerFactoryResults folder
    """
    import os
    import time

    date_string = time.strftime("%Y%m%d-%H%M%S")
    filename = 'Protection Study Results ' + date_string + ".xlsx"

    user = Path.home().name
    basepath = Path('//client/c$/LocalData') / user

    if basepath.exists():
        clientpath = basepath / Path('RelayCoordinationStudies')
    else:
        clientpath = Path('c:/LocalData') / user / Path('RelayCoordinationStudies')
    filepath = os.path.join(clientpath, filename)
    app.PrintPlain("Output file saved to " + filepath)

    feeder_name, study_type, grid_data_df, study_results, dfls_list, sect_trs = (
        format_results(study_type, gen_info, all_devices, setting_report, detailed_fls))

    #TODO: use Excel conditional formatting rules
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # General Information sheet
        grid_data_df.to_excel(writer, sheet_name='General Information', startrow=9, index=False)
        workbook = writer.book
        worksheet = workbook['General Information']
        worksheet['A1'] = feeder_name
        worksheet['A2'] = study_type
        worksheet['A4'] = 'Script Run Date'
        worksheet['A5'] = date_string
        worksheet['A8'] = 'External Grid Data:'

        # Study Results sheet
        study_results.to_excel(writer, sheet_name='Study Results', index=False)
        # Detailed Fault Levels sheet
        for i, device in enumerate(dfls_list):
            count = (i + 1) * 6 - 6
            device.to_excel(writer, sheet_name='Detailed Fault Levels', startrow=0, startcol=count, index=False)

    wb = load_workbook(filepath)
    ws = wb['General Information']
    adjust_col_width(ws)
    ws = wb['Study Results']
    adjust_col_width(ws)
    ws = wb['Detailed Fault Levels']
    adjust_col_width(ws)

    # Save the adjusted workbook
    wb.save(filepath)


def format_results(study_type, gen_info, all_devices, setting_report, detailed_fls):
    """

    :param study_type:
    :param gen_info:
    :param all_devices:
    :param setting_report:
    :param detailed_fls:
    :return:
    """

    # Format 'General Information' data
    instructs_lookup = {
        2: 'Full study (fault levels & relay coordination & grading diagram)',
        3: 'Fault level study only',
        4: 'Relay coordination study only',
        5: 'Create a grading diagram only',
        6: 'Line fuse study'
    }
    feeder_name = gen_info[0]
    study_type = instructs_lookup[study_type]
    grid_data = gen_info[1]
    grid_data_df = pd.DataFrame(grid_data)

    # Format 'Study Results' data
    formatted_dev = format_devices(all_devices)
    formatted_dev_pd = pd.DataFrame.from_dict(formatted_dev)
    setting_report_pd = pd.DataFrame.from_dict(setting_report)
    study_results = pd.concat([formatted_dev_pd, setting_report_pd])
    study_results = study_results.fillna("")

    # Format 'Detailed fault levels' data
    pg_max_all, phase_max_all, pg_min_all, phase_min_all, section_loads = detailed_fls

    def _loc_name_convert(dict):
        new_dic = {}
        for key, value in dict.items():
            new_dic[key.loc_name] = value
        return new_dic

    # Format section_loads data
    sect_trs = format_tfmrs(section_loads)

    device_list = [device for device in pg_max_all.keys()]
    dfls_list = []
    for device in device_list:
        dev_dict = {
            'Max PG fault': _loc_name_convert(pg_max_all[device]),
            'Max 3P fault': _loc_name_convert(phase_max_all[device]),
            'Min PG fault': _loc_name_convert(phase_max_all[device]),
            'Min 2P fault': _loc_name_convert(phase_max_all[device])
        }
        df = pd.DataFrame.from_dict(dev_dict, orient='index').transpose().reset_index()
        df = df.rename(columns={'index': device.loc_name})
        # Sort fault levels by Max PG fault
        df_sorted = df.sort_values(by=df.columns[1], ascending=False)
        df_sorted.insert(0, 'Tfmr Size (kVA)', '')

        # fill the transformer size column with terminal transformer size data.
        for df in sect_trs:
            if df.columns[0] == device.loc_name:
                target_df = df
                break
        tr_dict = pd.Series(target_df['Tfmr Size (kVA)'].values, index=target_df[device.loc_name]).to_dict()
        df_sorted['Tfmr Size (kVA)'] = df_sorted[device.loc_name].map(tr_dict).fillna('')

        dfls_list.append(df_sorted)

    return feeder_name, study_type, grid_data_df, study_results, dfls_list, sect_trs


def format_devices(all_devices: list) -> list[dict]:
    """

    :param all_devices:
    :return:
    """

    def check_att(dev, attribute):
        if hasattr(dev, attribute):
            return dev.attribute
        else:
            return ''

    device_list = {
        'Site Name':
            [
                'Voltage (kV)',
                'Current split n:1',
                'load (A)',
                'Rating  (A)',
                'DS capacity  (A)',
                'Max 3p FL',
                'Max PG FL',
                'Min 2P FL',
                'Min PG FL',
                'Max DS TR (Site name)',
                'Max TR size (kVA)',
                'Max TR fuse',
                'TR Max 3P ',
                'TR max PG',
                'DS devices',
                'BU device',
                'Device',
                'Status',
                'CT saturation',
                'CT composite error',
                'CT ratio',
                'OC pick up',
                'OC TMS',
                'OC Curve',
                'OC Hiset',
                'OC Min time (s)',
                'OC Hiset 2',
                'OC Min time 2 (s)',
                'EF pick up',
                'EF TMS',
                'EF Curve',
                'EF Hiset',
                'EF Min time (s)',
                'EF Hiset 2',
                'EF Min time 2 (s)'
             ]

    }

    for device in all_devices:
        device.netdat.downstream_devices = [device.name for device in device.netdat.downstream_devices]
        device.netdat.downstream_devices = ', '.join(device.netdat.downstream_devices)
        device.netdat.upstream_devices = [device.name for device in device.netdat.upstream_devices]
        device.netdat.upstream_devices = ', '.join(device.netdat.upstream_devices)
        device_list[device.name] = [
            device.netdat.voltage,
            device.netdat.i_split,
            device.netdat.load,
            device.netdat.rating,
            device.netdat.ds_capacity,
            device.netdat.max_3p_fl,
            device.netdat.max_pg_fl,
            device.netdat.min_2p_fl,
            device.netdat.min_pg_fl,
            device.netdat.tr_max_name,
            device.netdat.max_tr_size,
            device.netdat.max_tr_fuse,
            device.netdat.tr_max_3p,
            device.netdat.tr_max_pg,
            device.netdat.downstream_devices,
            device.netdat.upstream_devices,
            device.manufacturer.__name__,
            device.relset.status,
            check_att(device, 'ct.saturation'),
            check_att(device, 'ct.ect'),
            check_att(device, 'ct.ratio'),
            check_att(device, 'relset.oc_pu'),
            check_att(device, 'relset.oc_tms'),
            check_att(device, 'relset.oc_curve'),
            check_att(device, 'relset.oc_hiset'),
            check_att(device, 'relset.oc_min_time'),
            check_att(device, 'relset.oc_hiset2'),
            check_att(device, 'relset.oc_min_time2'),
            check_att(device, 'relset.ef_pu'),
            check_att(device, 'relset.ef_tms'),
            check_att(device, 'relset.ef_curve'),
            check_att(device, 'relset.ef_hiset'),
            check_att(device, 'relset.ef_min_time'),
            check_att(device, 'relset.ef_hiset2'),
            check_att(device, 'relset.ef_min_time2')
        ]

    return device_list


def format_tfmrs(section_loads):
    """

    :param section_loads:
    :return:
    """

    device_list = []
    for device, loads in section_loads.items():
        format_sect_loads = {
            device.loc_name: [elmlod.bus1.cterm.loc_name for elmlod in loads],
            'Tfmr Size (kVA)': [round(elmlod.Strat) for elmlod in loads]
        }
        format_sect_loads = pd.DataFrame.from_dict(format_sect_loads)
        device_list.append(format_sect_loads)
    return device_list


def adjust_col_width(ws):
    """
    Adjust column width of given Excel sheet
    :param ws:
    :return:
    """

    # Adjust the column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name

        # Iterate over all cells in the column to find the maximum length
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Set the column width to the maximum length found
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
